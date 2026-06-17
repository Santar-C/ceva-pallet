import os
import io
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from datetime import datetime
import qrcode
import psycopg2
import psycopg2.extras
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'super_secret_key_ceva_enterprise_v4')
DATABASE_URL = os.environ.get('DATABASE_URL')

def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    return conn

def dict_fetchall(cursor):
    columns = [desc[0] for desc in cursor.description]
    rows = []
    for row in cursor.fetchall():
        d = dict(zip(columns, row))
        for k, v in d.items():
            if isinstance(v, datetime):
                d[k] = v.strftime('%Y-%m-%d %H:%M:%S')
        rows.append(d)
    return rows

def dict_fetchone(cursor):
    columns = [desc[0] for desc in cursor.description]
    row = cursor.fetchone()
    if not row: return None
    d = dict(zip(columns, row))
    for k, v in d.items():
        if isinstance(v, datetime):
            d[k] = v.strftime('%Y-%m-%d %H:%M:%S')
    return d

def log_audit(action, detail, user):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("INSERT INTO audit_log (action, detail, user_name) VALUES (%s, %s, %s)", (action, detail, user))
        conn.commit(); cur.close(); conn.close()
    except Exception: pass

def init_db():
    conn = get_db(); cur = conn.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS transactions (
        id SERIAL PRIMARY KEY, tx_type TEXT NOT NULL, doc_number TEXT NOT NULL,
        country TEXT NOT NULL, po_number INTEGER NOT NULL, quantity INTEGER NOT NULL,
        base_qty INTEGER NOT NULL DEFAULT 0, lid_qty INTEGER NOT NULL DEFAULT 0,
        collar_qty INTEGER NOT NULL DEFAULT 0, user_name TEXT DEFAULT 'Unknown',
        timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    cur.execute("CREATE TABLE IF NOT EXISTS master_countries (id SERIAL PRIMARY KEY, name TEXT UNIQUE)")
    cur.execute('''CREATE TABLE IF NOT EXISTS audit_log (
        id SERIAL PRIMARY KEY, action TEXT, detail TEXT,
        user_name TEXT, timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY, username TEXT UNIQUE, password TEXT, role TEXT)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS stock_alerts (
        id SERIAL PRIMARY KEY, item_type TEXT UNIQUE, min_qty INTEGER DEFAULT 0)''')
    cur.execute("SELECT COUNT(*) FROM users")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO users (username, password, role) VALUES (%s, %s, %s)",
            ('admin', generate_password_hash('admin123'), 'admin'))
    cur.execute("SELECT COUNT(*) FROM master_countries")
    if cur.fetchone()[0] == 0:
        for c in ['Japan', 'Thailand', 'China', 'USA']:
            cur.execute("INSERT INTO master_countries (name) VALUES (%s) ON CONFLICT DO NOTHING", (c,))
    cur.execute("SELECT COUNT(*) FROM stock_alerts")
    if cur.fetchone()[0] == 0:
        for item in [('Base', 100), ('Lid', 100), ('Collar', 100)]:
            cur.execute("INSERT INTO stock_alerts (item_type, min_qty) VALUES (%s, %s) ON CONFLICT DO NOTHING", item)
    conn.commit(); cur.close(); conn.close()

def get_stock_data():
    conn = get_db(); cur = conn.cursor()
    cur.execute("""SELECT
        COALESCE(SUM(CASE WHEN tx_type='IN'  THEN base_qty   ELSE 0 END),0) AS in_base,
        COALESCE(SUM(CASE WHEN tx_type='OUT' THEN base_qty   ELSE 0 END),0) AS out_base,
        COALESCE(SUM(CASE WHEN tx_type='IN'  THEN lid_qty    ELSE 0 END),0) AS in_lid,
        COALESCE(SUM(CASE WHEN tx_type='OUT' THEN lid_qty    ELSE 0 END),0) AS out_lid,
        COALESCE(SUM(CASE WHEN tx_type='IN'  THEN collar_qty ELSE 0 END),0) AS in_collar,
        COALESCE(SUM(CASE WHEN tx_type='OUT' THEN collar_qty ELSE 0 END),0) AS out_collar
        FROM transactions""")
    row = dict_fetchone(cur); cur.close(); conn.close()
    row['total_base']   = row['in_base']   - row['out_base']
    row['total_lid']    = row['in_lid']    - row['out_lid']
    row['total_collar'] = row['in_collar'] - row['out_collar']
    return row

def get_alerts():
    stock = get_stock_data()
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM stock_alerts")
    alerts_cfg = dict_fetchall(cur); cur.close(); conn.close()
    alerts = []
    mapping = {'Base': 'total_base', 'Lid': 'total_lid', 'Collar': 'total_collar'}
    for a in alerts_cfg:
        current = stock.get(mapping.get(a['item_type'], ''), 0)
        if current <= a['min_qty']:
            alerts.append({'item': a['item_type'], 'current': current, 'min': a['min_qty']})
    return alerts

# ==================== AUTH ====================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = dict_fetchone(cur); cur.close(); conn.close()
        if user and check_password_hash(user['password'], password):
            session['username'] = user['username']
            session['role'] = user['role']
            log_audit('LOGIN', 'User logged in', username)
            return redirect(url_for('index'))
        flash("ชื่อผู้ใช้งาน หรือ รหัสผ่าน ไม่ถูกต้อง!", "error")
    return render_template('login.html')

@app.route('/logout')
def logout():
    log_audit('LOGOUT', 'User logged out', session.get('username', 'Unknown'))
    session.clear()
    return redirect(url_for('login'))

# ==================== USERS ====================
@app.route('/users')
def manage_users():
    if 'username' not in session: return redirect(url_for('login'))
    if session.get('role') != 'admin': flash("เฉพาะ Admin!", "error"); return redirect(url_for('index'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id, username, role FROM users ORDER BY role, username")
    users = dict_fetchall(cur); cur.close(); conn.close()
    return render_template('users.html', users=users)

@app.route('/add_user', methods=['POST'])
def add_user():
    if session.get('role') != 'admin': return redirect(url_for('index'))
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    role = request.form.get('role')
    if username and password:
        try:
            conn = get_db(); cur = conn.cursor()
            cur.execute("INSERT INTO users (username, password, role) VALUES (%s, %s, %s)",
                (username, generate_password_hash(password), role))
            conn.commit(); cur.close(); conn.close()
            log_audit('ADD_USER', f"Created: {username} ({role})", session['username'])
            flash(f"สร้างบัญชี '{username}' สำเร็จ!", "success")
        except psycopg2.IntegrityError:
            flash(f"ชื่อผู้ใช้ '{username}' มีอยู่แล้ว!", "error")
    return redirect(url_for('manage_users'))

@app.route('/delete_user/<int:user_id>')
def delete_user(user_id):
    if session.get('role') != 'admin': return redirect(url_for('index'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE id = %s", (user_id,))
    user = dict_fetchone(cur)
    if user['username'] == session['username']:
        flash("ไม่สามารถลบบัญชีตัวเองได้!", "error")
    else:
        cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
        conn.commit()
        log_audit('DELETE_USER', f"Deleted: {user['username']}", session['username'])
        flash(f"ลบบัญชี '{user['username']}' แล้ว", "success")
    cur.close(); conn.close()
    return redirect(url_for('manage_users'))

# ==================== MASTER DATA ====================
@app.route('/master_data')
def master_data():
    if 'username' not in session: return redirect(url_for('login'))
    if session.get('role') != 'admin': flash("เฉพาะ Admin!", "error"); return redirect(url_for('index'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM master_countries ORDER BY name")
    countries = dict_fetchall(cur)
    cur.execute("SELECT * FROM stock_alerts ORDER BY item_type")
    alerts = dict_fetchall(cur)
    cur.close(); conn.close()
    return render_template('master_data.html', countries=countries, alerts=alerts)

@app.route('/add_country', methods=['POST'])
def add_country():
    if session.get('role') != 'admin': return redirect(url_for('index'))
    name = request.form.get('country_name', '').strip()
    if name:
        try:
            conn = get_db(); cur = conn.cursor()
            cur.execute("INSERT INTO master_countries (name) VALUES (%s)", (name,))
            conn.commit(); cur.close(); conn.close()
            log_audit('ADD_MASTER', f"Added: {name}", session['username'])
            flash(f"เพิ่ม '{name}' สำเร็จ!", "success")
        except psycopg2.IntegrityError:
            flash(f"'{name}' มีอยู่แล้ว", "error")
    return redirect(url_for('master_data'))

@app.route('/delete_country/<int:c_id>')
def delete_country(c_id):
    if session.get('role') != 'admin': return redirect(url_for('index'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT name FROM master_countries WHERE id = %s", (c_id,))
    c = dict_fetchone(cur)
    if c:
        cur.execute("DELETE FROM master_countries WHERE id = %s", (c_id,))
        conn.commit()
        log_audit('DELETE_MASTER', f"Deleted: {c['name']}", session['username'])
        flash(f"ลบ '{c['name']}' แล้ว", "success")
    cur.close(); conn.close()
    return redirect(url_for('master_data'))

@app.route('/save_alerts', methods=['POST'])
def save_alerts():
    if session.get('role') != 'admin': return redirect(url_for('index'))
    conn = get_db(); cur = conn.cursor()
    for item in ['Base', 'Lid', 'Collar']:
        val = int(request.form.get(f'min_{item.lower()}', 0))
        cur.execute("UPDATE stock_alerts SET min_qty = %s WHERE item_type = %s", (val, item))
    conn.commit(); cur.close(); conn.close()
    flash("บันทึกค่าเตือนสต็อกแล้ว!", "success")
    return redirect(url_for('master_data'))

# ==================== MAIN INDEX ====================
@app.route('/')
def index():
    if 'username' not in session: return redirect(url_for('login'))
    start_date = request.args.get('start_date', '')
    end_date   = request.args.get('end_date', '')
    query  = "SELECT * FROM transactions WHERE 1=1"
    params = []
    if start_date: query += " AND DATE(timestamp) >= %s"; params.append(start_date)
    if end_date:   query += " AND DATE(timestamp) <= %s"; params.append(end_date)
    query += " ORDER BY timestamp DESC"
    conn = get_db(); cur = conn.cursor()
    cur.execute(query, params)
    log_rows = dict_fetchall(cur)
    cur.execute("SELECT name FROM master_countries ORDER BY name")
    countries = dict_fetchall(cur)
    cur.close(); conn.close()
    return render_template('index.html', stock_data=get_stock_data(), log_rows=log_rows,
        countries=countries, start_date=start_date, end_date=end_date, alerts=get_alerts())

# ==================== TRANSACTIONS ====================
@app.route('/process_transaction', methods=['POST'])
def process_transaction():
    if 'username' not in session: return redirect(url_for('login'))
    try:
        tx_type    = request.form['tx_type']
        doc_number = request.form['doc_number'].strip().upper()
        country    = request.form['country'].strip()
        po_number  = int(request.form.get('po_number', 0))
        quantity   = int(request.form.get('quantity', 0))
        user_name  = session['username']

        if tx_type == 'IN':
            if po_number == 0:
                # P00: รับเข้าเฉพาะ Base เท่านั้น
                calc_base   = quantity
                calc_lid    = 0
                calc_collar = 0
            else:
                # P01+: Base=Qty, Lid=Qty, Collar=PO×Qty
                calc_base   = quantity
                calc_lid    = quantity
                calc_collar = po_number * quantity
        else:
            # เบิกออก: กรอกแยกอิสระแต่ละชิ้น
            calc_base   = int(request.form.get('base_qty', 0))
            calc_lid    = int(request.form.get('lid_qty', 0))
            calc_collar = int(request.form.get('collar_qty', 0))
        if tx_type == 'OUT':
            current = get_stock_data()
            errors = []
            if calc_base   > current['total_base']:   errors.append(f"Base (มี {current['total_base']})")
            if calc_lid    > current['total_lid']:    errors.append(f"Lid (มี {current['total_lid']})")
            if calc_collar > current['total_collar']: errors.append(f"Collar (มี {current['total_collar']})")
            if errors:
                flash(f"สต็อกไม่พอ: {', '.join(errors)}", "error")
                return redirect(url_for('index'))

        tx_date = request.form.get('tx_date', '')
        try:
            tx_timestamp = datetime.strptime(tx_date, '%Y-%m-%dT%H:%M') if tx_date else datetime.now()
        except Exception:
            tx_timestamp = datetime.now()

        conn = get_db(); cur = conn.cursor()
        cur.execute("""INSERT INTO transactions
            (tx_type,doc_number,country,po_number,quantity,base_qty,lid_qty,collar_qty,user_name,timestamp)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (tx_type,doc_number,country,po_number,quantity,calc_base,calc_lid,calc_collar,user_name,tx_timestamp))
        conn.commit(); cur.close(); conn.close()
        log_audit('INSERT', f"{tx_type} Doc:{doc_number} B:{calc_base} L:{calc_lid} C:{calc_collar}", user_name)
        flash(f"บันทึก {'รับเข้า' if tx_type=='IN' else 'เบิกออก'} สำเร็จ!", "success")
    except Exception as e:
        flash(f"ข้อผิดพลาด: {e}", "error")
    return redirect(url_for('index'))

@app.route('/delete/<int:tx_id>')
def delete_tx(tx_id):
    if session.get('role') != 'admin': flash("เฉพาะ Admin!", "error"); return redirect(url_for('index'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM transactions WHERE id = %s", (tx_id,))
    tx = dict_fetchone(cur)
    cur.execute("DELETE FROM transactions WHERE id = %s", (tx_id,))
    conn.commit(); cur.close(); conn.close()
    log_audit('DELETE', f"Deleted ID:{tx_id} Doc:{tx['doc_number']}", session['username'])
    flash("ลบรายการสำเร็จ!", "success")
    return redirect(url_for('index'))

# ==================== ADMIN CLEAR DATA ====================
@app.route('/admin/clear', methods=['GET', 'POST'])
def admin_clear():
    if 'username' not in session: return redirect(url_for('login'))
    if session.get('role') != 'admin': flash("เฉพาะ Admin!", "error"); return redirect(url_for('index'))
    if request.method == 'POST':
        mode = request.form.get('clear_mode')
        conn = get_db(); cur = conn.cursor()
        count = 0
        if mode == 'all':
            cur.execute("SELECT COUNT(*) FROM transactions"); count = cur.fetchone()[0]
            cur.execute("DELETE FROM transactions")
            log_audit('CLEAR_ALL', f"Cleared ALL {count} records", session['username'])
            flash(f"ล้างข้อมูลทั้งหมด {count} รายการแล้ว!", "success")
        elif mode == 'by_date':
            start = request.form.get('start_date')
            end   = request.form.get('end_date')
            cur.execute("SELECT COUNT(*) FROM transactions WHERE DATE(timestamp) BETWEEN %s AND %s", (start, end))
            count = cur.fetchone()[0]
            cur.execute("DELETE FROM transactions WHERE DATE(timestamp) BETWEEN %s AND %s", (start, end))
            log_audit('CLEAR_DATE', f"Cleared {count} records {start}~{end}", session['username'])
            flash(f"ล้างข้อมูล {start} ถึง {end} จำนวน {count} รายการแล้ว!", "success")
        elif mode == 'keep_current_month':
            cur.execute("""SELECT COUNT(*) FROM transactions
                WHERE DATE_TRUNC('month', timestamp) != DATE_TRUNC('month', CURRENT_DATE)""")
            count = cur.fetchone()[0]
            cur.execute("""DELETE FROM transactions
                WHERE DATE_TRUNC('month', timestamp) != DATE_TRUNC('month', CURRENT_DATE)""")
            log_audit('CLEAR_KEEP_MONTH', f"Cleared {count} old records", session['username'])
            flash(f"ล้างข้อมูลเก่า {count} รายการ (เก็บเดือนนี้ไว้) แล้ว!", "success")
        conn.commit(); cur.close(); conn.close()
        return redirect(url_for('admin_clear'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM transactions"); total = cur.fetchone()[0]
    cur.execute("""SELECT COUNT(*) FROM transactions
        WHERE DATE_TRUNC('month', timestamp) = DATE_TRUNC('month', CURRENT_DATE)""")
    this_month = cur.fetchone()[0]
    cur.close(); conn.close()
    return render_template('admin_clear.html', total=total, this_month=this_month)

# ==================== DASHBOARD ====================
@app.route('/dashboard')
def dashboard():
    if 'username' not in session: return redirect(url_for('login'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("""SELECT TO_CHAR(timestamp,'YYYY-MM') AS month,
        SUM(CASE WHEN tx_type='IN'  THEN base_qty ELSE 0 END) AS in_base,
        SUM(CASE WHEN tx_type='OUT' THEN base_qty ELSE 0 END) AS out_base,
        SUM(CASE WHEN tx_type='IN'  THEN lid_qty  ELSE 0 END) AS in_lid,
        SUM(CASE WHEN tx_type='OUT' THEN lid_qty  ELSE 0 END) AS out_lid,
        SUM(CASE WHEN tx_type='IN'  THEN collar_qty ELSE 0 END) AS in_collar,
        SUM(CASE WHEN tx_type='OUT' THEN collar_qty ELSE 0 END) AS out_collar
        FROM transactions GROUP BY month ORDER BY month DESC LIMIT 6""")
    monthly = dict_fetchall(cur)
    monthly.reverse()
    cur.close(); conn.close()
    return render_template('dashboard.html', stock_data=get_stock_data(),
        monthly=monthly, alerts=get_alerts())

# ==================== AUDIT LOG ====================
@app.route('/audit_log')
def audit_log():
    if 'username' not in session: return redirect(url_for('login'))
    if session.get('role') != 'admin': flash("เฉพาะ Admin!", "error"); return redirect(url_for('index'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT 200")
    logs = dict_fetchall(cur); cur.close(); conn.close()
    return render_template('audit_log.html', logs=logs)


# ==================== EDIT TRANSACTION (Admin) ====================
@app.route('/edit/<int:tx_id>', methods=['GET', 'POST'])
def edit_tx(tx_id):
    if 'username' not in session: return redirect(url_for('login'))
    if session.get('role') != 'admin': flash("เฉพาะ Admin!", "error"); return redirect(url_for('index'))
    conn = get_db(); cur = conn.cursor()
    if request.method == 'POST':
        tx_type    = request.form['tx_type']
        doc_number = request.form['doc_number'].strip().upper()
        country    = request.form['country'].strip()
        po_number  = int(request.form.get('po_number', 0))
        quantity   = int(request.form.get('quantity', 0))
        base_qty   = int(request.form.get('base_qty', 0))
        lid_qty    = int(request.form.get('lid_qty', 0))
        collar_qty = int(request.form.get('collar_qty', 0))
        tx_date    = request.form.get('tx_date', '')
        try:
            tx_timestamp = datetime.strptime(tx_date, '%Y-%m-%dT%H:%M') if tx_date else datetime.now()
        except Exception:
            tx_timestamp = datetime.now()
        cur.execute("""UPDATE transactions SET
            tx_type=%s, doc_number=%s, country=%s, po_number=%s, quantity=%s,
            base_qty=%s, lid_qty=%s, collar_qty=%s, timestamp=%s
            WHERE id=%s""",
            (tx_type, doc_number, country, po_number, quantity,
             base_qty, lid_qty, collar_qty, tx_timestamp, tx_id))
        conn.commit()
        log_audit('EDIT', f"Edited ID:{tx_id} Doc:{doc_number}", session['username'])
        flash(f"แก้ไขรายการ {doc_number} สำเร็จ!", "success")
        cur.close(); conn.close()
        return redirect(url_for('index'))
    # GET
    cur.execute("SELECT * FROM transactions WHERE id = %s", (tx_id,))
    tx = dict_fetchone(cur)
    cur.execute("SELECT name FROM master_countries ORDER BY name")
    countries = dict_fetchall(cur)
    cur.close(); conn.close()
    if not tx: return "ไม่พบรายการ", 404
    return render_template('edit_tx.html', tx=tx, countries=countries)

# ==================== PRINT / QR ====================
@app.route('/print_slip/<doc_number>')
def print_slip(doc_number):
    if 'username' not in session: return redirect(url_for('login'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM transactions WHERE doc_number = %s", (doc_number,))
    items = dict_fetchall(cur); cur.close(); conn.close()
    if not items: return "ไม่พบเอกสาร", 404
    return render_template('slip.html', doc_number=doc_number, items=items,
        date=datetime.now().strftime('%d/%m/%Y %H:%M'), user=session['username'])

@app.route('/qr/<path:text>')
def generate_qr(text):
    img = qrcode.make(text); buf = io.BytesIO()
    img.save(buf, format='PNG'); buf.seek(0)
    return send_file(buf, mimetype='image/png')

# ==================== EXPORT ====================
@app.route('/export_excel')
def export_excel():
    if 'username' not in session: return redirect(url_for('login'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM transactions ORDER BY timestamp DESC")
    rows = dict_fetchall(cur); cur.close(); conn.close()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "รายการ"
    headers = ['#','ประเภท','เอกสาร','ประเทศ','PO','Set','Base','Lid','Collar','หมายเหตุ','ผู้ทำรายการ','เวลา']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="001A4C")
        cell.alignment = Alignment(horizontal='center')
    col_map = ['id','tx_type','doc_number','country','po_number','quantity','base_qty','lid_qty','collar_qty','note','user_name','timestamp']
    for i, row in enumerate(rows, 2):
        for j, k in enumerate(col_map, 1):
            v = row.get(k, '')
            if k == 'tx_type': v = 'รับเข้า' if v == 'IN' else 'เบิกออก'
            if k == 'timestamp' and v: v = str(v)[:16]
            ws.cell(row=i, column=j, value=v)
    output = io.BytesIO(); wb.save(output); output.seek(0)
    log_audit('EXPORT', f"Exported {len(rows)} records", session['username'])
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'CEVA_Pallet_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')



# ==================== IMPORT EXCEL ====================
@app.route('/import_excel', methods=['GET', 'POST'])
def import_excel():
    if 'username' not in session: return redirect(url_for('login'))
    if session.get('role') != 'admin': flash("เฉพาะ Admin!", "error"); return redirect(url_for('index'))
    if request.method == 'POST':
        f = request.files.get('excel_file')
        if not f or not f.filename.endswith(('.xlsx','.xls')):
            flash("กรุณาเลือกไฟล์ Excel (.xlsx/.xls)", "error")
            return redirect(url_for('import_excel'))
        try:
            from openpyxl import load_workbook
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col = {h: i for i, h in enumerate(headers)}
            required = ['ประเภท','เอกสาร','ประเทศ','PO','Set','Base','Lid','Collar','ผู้ทำรายการ','เวลา']
            missing = [r for r in required if r not in col]
            if missing:
                flash(f"ไม่พบคอลัมน์: {', '.join(missing)}", "error")
                return redirect(url_for('import_excel'))
            conn = get_db(); cur = conn.cursor()
            count = 0
            errors = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    tx_type    = 'IN' if str(row[col['ประเภท']]) in ['IN','รับเข้า'] else 'OUT'
                    doc_number = str(row[col['เอกสาร']] or '').strip().upper()
                    country    = str(row[col['ประเทศ']] or '').strip()
                    po_number  = int(row[col['PO']] or 0)
                    quantity   = int(row[col['Set']] or 0)
                    base_qty   = int(row[col['Base']] or 0)
                    lid_qty    = int(row[col['Lid']] or 0)
                    collar_qty = int(row[col['Collar']] or 0)
                    user_name  = str(row[col['ผู้ทำรายการ']] or session['username']).strip()
                    ts_raw     = row[col['เวลา']]
                    if isinstance(ts_raw, datetime):
                        ts = ts_raw
                    elif ts_raw:
                        ts = datetime.strptime(str(ts_raw)[:16], '%Y-%m-%d %H:%M')
                    else:
                        ts = datetime.now()
                    if not doc_number or not country: continue
                    cur.execute("""INSERT INTO transactions
                        (tx_type,doc_number,country,po_number,quantity,base_qty,lid_qty,collar_qty,user_name,timestamp)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (tx_type,doc_number,country,po_number,quantity,base_qty,lid_qty,collar_qty,user_name,ts))
                    count += 1
                except Exception as e:
                    errors.append(f"แถว {i}: {e}")
            conn.commit(); cur.close(); conn.close()
            log_audit('IMPORT', f"Imported {count} records from Excel", session['username'])
            msg = f"นำเข้าสำเร็จ {count} รายการ"
            if errors: msg += f" (ข้ามไป {len(errors)} แถว)"
            flash(msg, "success")
        except Exception as e:
            flash(f"ข้อผิดพลาด: {e}", "error")
        return redirect(url_for('import_excel'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM transactions"); total = cur.fetchone()[0]
    cur.close(); conn.close()
    return render_template('import_excel.html', total=total)

# ==================== REPORTS ====================
@app.route('/reports')
def reports():
    if 'username' not in session: return redirect(url_for('login'))
    conn = get_db(); cur = conn.cursor()

    # Monthly summary
    cur.execute("""
        SELECT TO_CHAR(timestamp,'YYYY-MM') AS month,
            SUM(CASE WHEN tx_type='IN'  THEN base_qty   ELSE 0 END) AS in_base,
            SUM(CASE WHEN tx_type='OUT' THEN base_qty   ELSE 0 END) AS out_base,
            SUM(CASE WHEN tx_type='IN'  THEN lid_qty    ELSE 0 END) AS in_lid,
            SUM(CASE WHEN tx_type='OUT' THEN lid_qty    ELSE 0 END) AS out_lid,
            SUM(CASE WHEN tx_type='IN'  THEN collar_qty ELSE 0 END) AS in_collar,
            SUM(CASE WHEN tx_type='OUT' THEN collar_qty ELSE 0 END) AS out_collar,
            COUNT(*) AS tx_count
        FROM transactions GROUP BY month ORDER BY month DESC LIMIT 12""")
    monthly = dict_fetchall(cur)

    # Top countries OUT
    cur.execute("""
        SELECT country,
            SUM(base_qty) AS base, SUM(lid_qty) AS lid, SUM(collar_qty) AS collar,
            COUNT(*) AS tx_count
        FROM transactions WHERE tx_type='OUT'
        GROUP BY country ORDER BY base DESC LIMIT 10""")
    top_countries = dict_fetchall(cur)

    # Forecast: avg daily OUT last 30 days
    cur.execute("""
        SELECT
            ROUND(SUM(base_qty)::numeric   / GREATEST(COUNT(DISTINCT DATE(timestamp)),1), 2) AS avg_base,
            ROUND(SUM(lid_qty)::numeric    / GREATEST(COUNT(DISTINCT DATE(timestamp)),1), 2) AS avg_lid,
            ROUND(SUM(collar_qty)::numeric / GREATEST(COUNT(DISTINCT DATE(timestamp)),1), 2) AS avg_collar
        FROM transactions
        WHERE tx_type='OUT' AND timestamp >= NOW() - INTERVAL '30 days'""")
    avg30 = dict_fetchone(cur) or {'avg_base':0,'avg_lid':0,'avg_collar':0}

    # Forecast: trend (last 30 vs prev 30)
    cur.execute("""
        SELECT
            ROUND(SUM(CASE WHEN timestamp >= NOW()-INTERVAL'30 days' THEN base_qty   ELSE 0 END)::numeric/30,2) AS new_base,
            ROUND(SUM(CASE WHEN timestamp <  NOW()-INTERVAL'30 days' AND timestamp >= NOW()-INTERVAL'60 days' THEN base_qty ELSE 0 END)::numeric/30,2) AS old_base,
            ROUND(SUM(CASE WHEN timestamp >= NOW()-INTERVAL'30 days' THEN lid_qty    ELSE 0 END)::numeric/30,2) AS new_lid,
            ROUND(SUM(CASE WHEN timestamp <  NOW()-INTERVAL'30 days' AND timestamp >= NOW()-INTERVAL'60 days' THEN lid_qty  ELSE 0 END)::numeric/30,2) AS old_lid,
            ROUND(SUM(CASE WHEN timestamp >= NOW()-INTERVAL'30 days' THEN collar_qty ELSE 0 END)::numeric/30,2) AS new_collar,
            ROUND(SUM(CASE WHEN timestamp <  NOW()-INTERVAL'30 days' AND timestamp >= NOW()-INTERVAL'60 days' THEN collar_qty ELSE 0 END)::numeric/30,2) AS old_collar
        FROM transactions WHERE tx_type='OUT'""")
    trend = dict_fetchone(cur) or {}

    stock = get_stock_data()
    cur.close(); conn.close()

    def days_left(stock_val, avg):
        try:
            avg = float(avg) if avg else 0
            return round(float(stock_val) / avg) if avg > 0 else 999
        except: return 999

    forecast = {
        'base_days':   days_left(stock['total_base'],   avg30['avg_base']),
        'lid_days':    days_left(stock['total_lid'],    avg30['avg_lid']),
        'collar_days': days_left(stock['total_collar'], avg30['avg_collar']),
        'avg30': avg30, 'trend': trend
    }

    return render_template('reports.html',
        monthly=monthly, top_countries=top_countries,
        forecast=forecast, stock=stock, alerts=get_alerts())

# ==================== KPI ====================
@app.route('/kpi')
def kpi():
    if 'username' not in session: return redirect(url_for('login'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) AS total_tx,
            SUM(CASE WHEN tx_type='IN'  THEN 1 ELSE 0 END) AS total_in,
            SUM(CASE WHEN tx_type='OUT' THEN 1 ELSE 0 END) AS total_out,
            SUM(CASE WHEN DATE_TRUNC('month',timestamp)=DATE_TRUNC('month',NOW()) THEN 1 ELSE 0 END) AS this_month,
            COUNT(DISTINCT user_name) AS active_users,
            COUNT(DISTINCT country) AS countries_served
        FROM transactions""")
    kpi_data = dict_fetchone(cur)
    cur.execute("""
        SELECT TO_CHAR(timestamp,'YYYY-MM-DD') AS day, COUNT(*) AS cnt
        FROM transactions
        WHERE timestamp >= NOW() - INTERVAL '14 days'
        GROUP BY day ORDER BY day""")
    daily = dict_fetchall(cur)
    cur.execute("""
        SELECT user_name, COUNT(*) AS cnt
        FROM transactions GROUP BY user_name ORDER BY cnt DESC LIMIT 5""")
    top_users = dict_fetchall(cur)
    cur.close(); conn.close()
    return render_template('kpi.html', kpi=kpi_data, daily=daily,
        top_users=top_users, stock=get_stock_data(), alerts=get_alerts())

# ==================== AUTO LOGOUT CONFIG ====================
@app.route('/set_session_timeout', methods=['POST'])
def set_session_timeout():
    if session.get('role') != 'admin': return redirect(url_for('index'))
    minutes = int(request.form.get('timeout_minutes', 30))
    session['timeout_minutes'] = minutes
    flash(f"ตั้งค่า Auto Logout {minutes} นาทีแล้ว", "success")
    return redirect(url_for('manage_users'))

with app.app_context():
    init_db()

if __name__ == '__main__':
    app.run(debug=False)
